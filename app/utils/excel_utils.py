"""Excel utilities — replaces ExcelUtils.java using openpyxl."""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

log = logging.getLogger(__name__)

MAX_UNLIMIT = -1


@dataclass
class Cell:
    """A single spreadsheet cell value."""
    value: Any
    row: int = 0
    col: int = 0

    def as_string(self) -> str:
        return str(self.value) if self.value is not None else ""


@dataclass
class IRow:
    """A spreadsheet row with its cells and row number."""
    cells: list[Cell] = field(default_factory=list)
    row_no: int = 0


def read_excel(path: str | Path, max_rows: int = MAX_UNLIMIT, has_head: bool = True) -> list[list[Cell]]:
    """Read an Excel file and return rows of ``Cell`` objects."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[Cell]] = []
    row_no = 0

    for row in ws.iter_rows(values_only=True):
        if has_head and row_no == 0:
            row_no += 1
            continue
        if max_rows > 0 and len(rows) >= max_rows:
            break
        cells = [Cell(value=v, row=row_no, col=i) for i, v in enumerate(row)]
        rows.append(cells)
        row_no += 1

    wb.close()
    return rows


def read_excel_rows(path: str | Path, sheet_no: int = 0) -> list[IRow]:
    """Read an Excel file and return ``IRow`` objects."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_no]
    rows: list[IRow] = []
    row_no = 0

    for row in ws.iter_rows(values_only=True):
        cells = [Cell(value=v, row=row_no, col=i) for i, v in enumerate(row)]
        rows.append(IRow(cells=cells, row_no=row_no))
        row_no += 1

    wb.close()
    return rows


def re_save_and_calc_formula(path: str | Path, calc_formula: bool = True) -> bool:
    """Open and re-save an Excel file so formulas are evaluated."""
    path = Path(path)
    try:
        wb = load_workbook(path)
        if calc_formula:
            # openpyxl recalculates on next Excel open; we just save
            pass
        wb.save(path)
        wb.close()
        return True
    except Exception as exc:
        log.error("Re-save excel error : %s : %s", path, exc)
        return False


def save_to_csv(path: str | Path, encoding: str = "utf-8") -> Path:
    """Convert an Excel file to CSV and return the new path."""
    path = Path(path)
    csv_path = path.with_suffix(".csv")

    rows = read_excel(path)
    with open(csv_path, "w", newline="", encoding=encoding) as f:
        if encoding.lower() == "utf-8":
            f.write("\ufeff")  # UTF-8 BOM for Excel compatibility
        writer = csv.writer(f)
        for row in rows:
            writer.writerow([cell.as_string() for cell in row])

    return csv_path
